import openpyxl  # Import the openpyxl library to work with Excel files


# Function to extract data from an Excel sheet and return it in a list format
def get_data(sheetName):
    # Load the Excel workbook from the specified file path
    workbook = openpyxl.load_workbook("..//excel//testdata.xlsx")
    # Select the sheet from which data needs to be retrieved
    sheet = workbook[sheetName]

    # Get the total number of rows and columns in the sheet
    totalrows = sheet.max_row
    totalcols = sheet.max_column

    # Print the total number of columns and rows (for debugging or logging purposes)
    print("Total cols are ", str(totalcols))
    print("Total rows are ", str(totalrows))

    # Create an empty list to hold the main set of data (list of lists)
    mainList = []

    # Loop through each row starting from the second row (usually skipping headers)
    for i in range(2, totalrows + 1):  # Loop through all rows except the header (row 1)
        dataList = []  # Temporary list to hold data from the current row

        # Loop through each column to get data from each cell in the current row
        for j in range(1, totalcols + 1):  # Loop through all columns in the row
            # Get the value from the specific cell (row=i, column=j)
            data = sheet.cell(row=i, column=j).value
            # Insert the retrieved data into the temporary list (dataList)
            dataList.insert(j, data)

        # Once the row data is collected, insert the dataList into the mainList
        mainList.insert(i, dataList)
        # Print the main list to see the updated data after each row is processed (for debugging)
        print(mainList)

    # Return the main list containing all the rows of data
    return mainList
