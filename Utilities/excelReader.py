import openpyxl  # Importing the openpyxl library to work with Excel files

# Function to get the total number of rows in a given Excel sheet
def getRowCount(path, sheetName):
    # Load the Excel workbook from the specified path
    workbook = openpyxl.load_workbook(path)
    # Select the sheet by its name
    sheet = workbook[sheetName]
    # Return the maximum number of rows in the sheet
    return sheet.max_row

# Function to get the total number of columns in a given Excel sheet
def getColCount(path, sheetName):
    # Load the Excel workbook from the specified path
    workbook = openpyxl.load_workbook(path)
    # Select the sheet by its name
    sheet = workbook[sheetName]
    # Return the maximum number of columns in the sheet
    return sheet.max_column

# Function to get the data from a specific cell in the Excel sheet
def getCellData(path, sheetName, rowNum, colNum):
    # Load the Excel workbook from the specified path
    workbook = openpyxl.load_workbook(path)
    # Select the sheet by its name
    sheet = workbook[sheetName]
    # Retrieve the value from the specified cell (rowNum and colNum)
    return sheet.cell(row=rowNum, column=colNum).value

# Function to set data into a specific cell in the Excel sheet
def setCellData(path, sheetName, rowNum, colNum, data):
    # Load the Excel workbook from the specified path
    workbook = openpyxl.load_workbook(path)
    # Select the sheet by its name
    sheet = workbook[sheetName]
    # Set the specified data into the cell at rowNum and colNum
    sheet.cell(row=rowNum, column=colNum).value = data
    # Save the workbook after modifying the cell data
    workbook.save(path)

# Define the path to the Excel file and the name of the sheet to work with
path = "..//excel//testdata.xlsx"
sheetName = "LoginTest"

# Get the total number of rows and columns in the sheet
rows = getRowCount(path, sheetName)
cols = getColCount(path, sheetName)

# Print the total number of rows and columns
print(rows, "---", cols)

# Get the data from the cell located at row 2, column 1, and print it
print(getCellData(path, sheetName, 2, 1))

# Set the value "DOB" into the cell located at row 1, column 4, and save the workbook
setCellData(path, sheetName, 1, 4, "DOB")
